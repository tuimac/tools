import openpyxl
import glob
import gzip
import os
import datetime

LOG_DIR = '03'
REPORT_FILENAME = 'report.xlsx'

def importToWorkbook(sheet, line, timestamp, cellindex):
    if cellindex == 1:
        sheet.cell(row=cellindex, column=1).value = 'timestamp'
    else:
        sheet.cell(row=cellindex, column=1).value = datetime.datetime.strptime(timestamp, '%Y%m%dT%H%MZ')
        sheet['A' + str(cellindex)].number_format = 'yyyy/mm/dd hh:mm'
    index = 2
    for element in line.split(' '):
        sheet.cell(row=cellindex, column=index).value = element
        index += 1

if __name__ == '__main__':
    workbook = openpyxl.Workbook()
    sheet = workbook.create_sheet('result')
    cellindex = 1
    os.chdir(LOG_DIR)

    for filename in glob.glob(r'*.gz'):
        with gzip.open(filename, 'rb') as logfile:
            lineindex = 0
            for line in logfile.readlines():
                if lineindex == 0:
                    if cellindex != 1:
                        continue
                importToWorkbook(sheet, line.decode(), filename.split('_')[4], cellindex)
                cellindex += 1
                lineindex += 1
    
    workbook.save(REPORT_FILENAME)
    del workbook['Sheet']
    workbook.save(REPORT_FILENAME)
