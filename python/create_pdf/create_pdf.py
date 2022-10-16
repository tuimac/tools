import openpyxl
import json
import time
import os
import sys
import stat
import traceback
import platform
import zipfile
import urllib.request
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

CHROME_VERSION = '106.0.5249.119'
EXCEL_NAME = 'pdf.xlsx'
EXCEL_SHEET_NAME = 'シート1'
DRIVER_DIR = '.'
PDF_DIR = '.'

DRIVER_MAP = {
    'Linux': 'chromedriver_linux64.zip',
    'Darwin': 'chromedriver_mac64.zip',
    'Windows': 'chromedriver_win32.zip'
}

def import_excel() -> dict:
    data = dict()
    workbook = openpyxl.load_workbook(EXCEL_NAME)
    sheet = workbook[EXCEL_SHEET_NAME]
    row = 2
    
    while True:
        id_num = sheet.cell(row=row, column=1).value
        if id_num == None:
            break
        else:
            hello_work_num = sheet.cell(row=row, column=2).value
            url = 'https://www.hellowork.mhlw.go.jp/kensaku/GECA110010.do?screenId=GECA110010&action=dispDetailBtn&kJNo=' + hello_work_num + '&kJKbn=1&jGSHNo=B5VtkNVFHYHEOTF5MdwAqQ%3D%3D&fullPart=2&iNFTeikyoRiyoDtiID=&kSNo=&newArrived=&tatZngy=1&shogaiKbn=0'
            data[id_num] = url
            row += 1
    return data

def __check_chromedriver_path(os_type) -> str:
    if os_type == 'Linux' or 'Darwin':
        driver_file_path = DRIVER_DIR+ '/chromedriver'
        if os.path.exists(driver_file_path) is True:
            return driver_file_path
        else:
            return ''
    elif os_type == 'Windows':
        driver_file_path = DRIVER_DIR + '/chromedriver.exe'
        if os.path.exists(driver_file_path) is True:
            return driver_file_path
        else:
            return ''
    else:
        raise KeyError

def download_chromedriver() -> str:
    try:
        os_type = platform.system()
        
        # Confirm if there is Chrome driver under the DRIVER_DIR
        chromedriver_path = __check_chromedriver_path(os_type)
        if chromedriver_path != '':
            return chromedriver_path

        # Get the latest version information for Chrome on this machine
        main_chrome_version = CHROME_VERSION.split('.')[0]
        handler = urllib.request.urlopen(
            'https://chromedriver.storage.googleapis.com/LATEST_RELEASE_' + main_chrome_version
        )
        latest_driver_version = handler.read().decode()
        handler.close()
        
        # Download the valid version Chrome driver for this machine
        download_path = DRIVER_DIR + '/' + DRIVER_MAP[os_type]
        handler = urllib.request.urlretrieve(
            'https://chromedriver.storage.googleapis.com/' + latest_driver_version + '/' + DRIVER_MAP[os_type],
            download_path 
        )
        with zipfile.ZipFile(download_path) as zip_handler:
            zip_handler.extractall(DRIVER_DIR)
        os.remove(download_path)
    
        # Confirm if there is Chrome driver under the DRIVER_DIR
        chromedriver_path = __check_chromedriver_path(os_type)
        print(chromedriver_path)
        if chromedriver_path == '':
            raise KeyError
        else:
            os.chmod(chromedriver_path, 0o755)
            return chromedriver_path

    except KeyError as e:
        print('There is no support for this OS.(' + os_type + ')', file=sys.stderr)
        os._exit(1)
    except:
        traceback.print_exc()
        os._exit(1)

def create_pdf(data, driver_path):
       
    # Setup Chrome options for prinr page as PDF
    chrome_option = webdriver.ChromeOptions()
    printer_config = {
        'recentDestinations': [
            {
                'id': 'Save as PDF',
                'origin': 'local',
                'account':''
            }
        ],
        'selectedDestinationId': 'Save as PDF',
        'version': 2,
        'isLandscapeEnabled': False,
        'pageSize': 'A4',
        #'mediaSize': {'height_microns': 355600, 'width_microns': 215900}, #紙のサイズ　（10000マイクロメートル = １cm）
        'marginsType': 0,
        'scalingType': 0,
        #'scaling': '141' ,#倍率
        'isHeaderFooterEnabled': False,
        'isCssBackgroundEnabled': True,
        'isDuplexEnabled': False,
        'isColorEnabled': True,
        'isCollateEnabled': True
    }
    
    prefs = {
        'printing.print_preview_sticky_settings.appState': json.dumps(printer_config),
        'download.default_directory': os.getcwd(),
        'plugins.always_open_pdf_externally': True,
        'download.prompt_for_download': False,
        'download.directory_upgrade': True
    }

    chrome_option.add_experimental_option('prefs', prefs)
    chrome_option.add_argument('--kiosk-printing')

    # Create PDF from each url website
    driver = webdriver.Chrome(executable_path=driver_path, options=chrome_option)
    for key in data:
        url = data[key]
        driver.implicitly_wait(10)
        driver.get(url)
        WebDriverWait(driver, 15).until(EC.presence_of_all_elements_located)
        driver.execute_script('document.title="' + key + '";window.print();')
        time.sleep(1)
    driver.quit()

if __name__ == '__main__':
    try:
        data = import_excel()
        driver_path = download_chromedriver()
        create_pdf(data, driver_path)
    except:
        traceback.print_exc()
